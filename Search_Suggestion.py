import openpyxl
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from webdriver_manager.chrome import ChromeDriverManager

# Function to extract keywords from the specified column and sheet
def extract_keywords(workbook, sheet_name, column_number):
    sheet = workbook[sheet_name]
    keywords = []
    for row in sheet.iter_rows(min_row=3, values_only=True):
        keyword = row[column_number - 1]
        if keyword:
            keywords.append(keyword)
    return keywords

# Function to search keywords on Google
def search_keywords(driver, keywords):
    for keyword in keywords:
        search_box = driver.find_element(By.NAME, "q")
        search_box.clear()
        search_box.send_keys(keyword)
        time.sleep(2)  # Wait for suggestions to load

        suggestions = driver.find_elements(By.CSS_SELECTOR, 'li.sbct span')
        options = [suggestion.text for suggestion in suggestions if suggestion.text]
        if options:
            longest_option = max(options, key=len)
            shortest_option = min(options, key=len)
        print(longest_option +" , "+ shortest_option)

# Main function
def main():
    # Replace '4BeatsQ1.xlsx' with the actual file path
    workbook = openpyxl.load_workbook('F:/#####/Automation_Using_Selenium/4BeatsQ1.xlsx')

    # Get the current sheet name (today's date)
    workbook_path = 'F:/#####/Automation_Using_Selenium/4BeatsQ1.xlsx'
    workbook = openpyxl.load_workbook(workbook_path)
    current_day = datetime.now().strftime('%A')  # e.g., 'Monday'

    # Extract keywords from column C[3]
    keywords = extract_keywords(workbook, current_day, 3)
    print(keywords)

    # Step 1: Set the path to your Chrome WebDriver (if not already in PATH)
    driver = webdriver.Chrome()
    # Step 2: Maximize the browser window (optional)
    driver.maximize_window()
    # Step 3: Navigate to the Google home page
    driver.get("https://www.google.com/")

    # Search keywords on Google
    search_keywords(driver, keywords)
    # Close the browser
    driver.quit()

if __name__ == "__main__":
    main()