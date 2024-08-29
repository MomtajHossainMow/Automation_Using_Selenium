import openpyxl
import time
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager

# Function to perform Google search and return longest and shortest suggestions
def search_google(driver, keyword):
    try:
        search_box = driver.find_element(By.NAME, "q")  #Find the search box
        search_box.clear()                              #Clear the search box
        search_box.send_keys(keyword)                   #Write the keyword
        time.sleep(2)  # Wait for suggestions to load
        #Collect the options in suggestion
        suggestions = driver.find_elements(By.CSS_SELECTOR, 'li.sbct span')
        options = [suggestion.text for suggestion in suggestions if suggestion.text]
        if options:
            longest_option = max(options, key=len)
            shortest_option = min(options, key=len)
            return longest_option, shortest_option
    except Exception as e:
        print(f"Error searching for keyword '{keyword}': {e}")
    return None, None




# Load the workbook and select the sheet based on the current day
workbook_path = 'F:/#####/Automation_Using_Selenium/Q1_Assignment/4BeatsQ1.xlsx'
workbook = openpyxl.load_workbook(workbook_path)
current_day = datetime.now().strftime('%A')  # e.g., 'Friday'
sheet = workbook[current_day]

# Read keywords from column C of the sheet
keywords = []
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
    if row[0] is not None:  # Ensure keyword is not None
        keywords.append(row[0])  # Keywords are in column C

# Initialize ChromeDriver
driver = webdriver.Chrome()
# Navigate to the Google home page
driver.get("https://www.google.com/")
# Dictionary to store results
results = {}
# Call the search_google() for each keyword and store the result
for keyword in keywords:
    longest, shortest = search_google(driver, keyword)
    results[keyword] = {'longest': longest, 'shortest': shortest}
    print(f"Keyword: {keyword}, Longest: {longest}, Shortest: {shortest}")  # Debug statement
# Close the browser
driver.quit()

# Write results back to the Excel sheet in columns D (Longest Option) and E (Shortest Option)
for row in sheet.iter_rows(min_row=2, min_col=3, max_col=5):
    keyword = row[0].value
    if keyword in results:
        sheet.cell(row=row[0].row, column=4).value = results[keyword]['longest']  # Longest option in column D
        sheet.cell(row=row[0].row, column=5).value = results[keyword]['shortest']  # Shortest option in column E

# Save the workbook
output_path = 'F:/#####/Automation_Using_Selenium/Q1_Assignment/4BeatsQ1.xlsx' # Updated save location
workbook.save(output_path)
print(f"Results saved to {output_path}")
